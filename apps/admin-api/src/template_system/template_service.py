from __future__ import annotations

import re
import unicodedata
from datetime import datetime, timedelta
from io import BytesIO
from typing import Any, Dict, List, Optional, Tuple

from sqlalchemy import distinct, func
from sqlalchemy.orm import Session
from PIL import Image

from .preview_engine import PreviewConfig, PreviewEngine, PreviewFormat, PreviewQuality
from .models import PlantillaAnalitica, PlantillaRutina
from .template_repository import TemplateRepository
from .template_validator import ValidationResult


class TemplateService:
    def __init__(self, db_session: Session):
        self.db = db_session
        self.repository = TemplateRepository(db_session)
        self.preview_engine = PreviewEngine()

    def _normalize_metadata_name(self, configuracion: Dict[str, Any]) -> Dict[str, Any]:
        if not isinstance(configuracion, dict):
            return configuracion
        metadata = configuracion.get("metadata")
        if not isinstance(metadata, dict):
            return configuracion
        name = str(metadata.get("name") or "").strip()
        if not name:
            return configuracion
        normalized = unicodedata.normalize("NFKD", name)
        normalized = "".join(c for c in normalized if not unicodedata.combining(c))
        normalized = re.sub(r"[^a-zA-Z0-9_\-\s]+", "", normalized)
        normalized = re.sub(r"\s+", " ", normalized).strip()
        if normalized:
            metadata["name"] = normalized
        return configuracion

    def create_template(self, payload: Dict[str, Any], creada_por: Optional[int] = None) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        nombre = str(payload.get("nombre") or "").strip()
        if not nombre:
            return None, "nombre requerido"
        configuracion = payload.get("configuracion")
        if not isinstance(configuracion, dict):
            return None, "configuracion requerida"
        configuracion = self._normalize_metadata_name(configuracion)
        descripcion = payload.get("descripcion")
        categoria = str(payload.get("categoria") or "general")
        dias_semana = payload.get("dias_semana")
        try:
            dias_semana = int(dias_semana) if dias_semana is not None else None
        except Exception:
            dias_semana = None
        tags = payload.get("tags")
        if tags is not None and not isinstance(tags, list):
            tags = None
        publica = bool(payload.get("publica") or False)
        activa = bool(payload.get("activa") if payload.get("activa") is not None else True)

        tpl, err = self.repository.create_template(
            nombre=nombre,
            configuracion=configuracion,
            descripcion=str(descripcion) if descripcion else None,
            categoria=categoria,
            dias_semana=dias_semana,
            creada_por=creada_por,
            tags=[str(t) for t in tags] if isinstance(tags, list) else None,
            publica=publica,
            activa=activa,
        )
        if not tpl:
            return None, err or "create_failed"
        return self._template_to_dict(tpl), None

    def get_template(self, template_id: int) -> Optional[Dict[str, Any]]:
        tpl = self.repository.get_template(int(template_id))
        if not tpl:
            return None
        return self._template_to_dict(tpl)

    def update_template(self, template_id: int, payload: Dict[str, Any], creada_por: Optional[int] = None, cambios_descripcion: Optional[str] = None) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        allowed = {"nombre", "configuracion", "descripcion", "categoria", "dias_semana", "tags", "publica", "activa", "preview_url"}
        updates: Dict[str, Any] = {}
        for k, v in payload.items():
            if k in allowed:
                updates[k] = v
        if "nombre" in updates:
            updates["nombre"] = str(updates["nombre"] or "").strip()
        if "categoria" in updates:
            updates["categoria"] = str(updates["categoria"] or "general")
        if "dias_semana" in updates:
            try:
                updates["dias_semana"] = int(updates["dias_semana"]) if updates["dias_semana"] is not None else None
            except Exception:
                updates["dias_semana"] = None
        if "tags" in updates and updates["tags"] is not None and not isinstance(updates["tags"], list):
            updates["tags"] = None
        if "configuracion" in updates and isinstance(updates["configuracion"], dict):
            updates["configuracion"] = self._normalize_metadata_name(updates["configuracion"])

        tpl, err = self.repository.update_template(int(template_id), updates=updates, creada_por=creada_por, cambios_descripcion=cambios_descripcion)
        if not tpl:
            return None, err or "update_failed"
        return self._template_to_dict(tpl), None

    def delete_template(self, template_id: int) -> Tuple[bool, Optional[str]]:
        return self.repository.delete_template(int(template_id))

    def search_templates(self, params: Dict[str, Any]) -> Dict[str, Any]:
        query = params.get("query")
        categoria = params.get("categoria")
        activa = params.get("activa")
        if activa is not None:
            activa = True if str(activa).lower() == "true" else False if str(activa).lower() == "false" else None
        sort_by = str(params.get("sort_by") or "fecha_creacion")
        sort_order = str(params.get("sort_order") or "desc")
        limit = int(params.get("limit") or 50)
        offset = int(params.get("offset") or 0)

        templates, total = self.repository.search_templates(
            query=str(query) if query else None,
            categoria=str(categoria) if categoria else None,
            activa=activa,
            sort_by=sort_by,
            sort_order=sort_order,
            limit=limit,
            offset=offset,
            export_only=True,
        )
        has_more = offset + limit < total
        return {
            "success": True,
            "templates": [self._template_to_dict(t) for t in templates],
            "total": total,
            "has_more": has_more,
            "limit": limit,
            "offset": offset,
        }

    def validate_template_config(self, configuracion: Dict[str, Any]) -> ValidationResult:
        return self.repository.validate_template(configuracion)

    def generate_template_preview(self, template_config: Dict[str, Any], format: str, quality: str, page_number: int, sample_data: Optional[Dict[str, Any]]) -> Tuple[Optional[str], Optional[str]]:
        try:
            pf = PreviewFormat(format.lower())
        except Exception:
            return None, "Invalid format"
        try:
            pq = PreviewQuality(quality.lower())
        except Exception:
            return None, "Invalid quality"

        cfg = PreviewConfig(format=pf, quality=pq, page_number=int(page_number or 1), use_cache=True, generate_sample_data=sample_data is None)
        result = self.preview_engine.generate_preview(template_config=template_config, config=cfg, custom_data=sample_data)
        if not result.success:
            return None, result.error_message or "preview_failed"
        uri = self.preview_engine.build_data_uri(result)
        if not uri:
            return None, "preview_failed"
        return uri, None

    def generate_and_store_thumbnail(self, template_id: int) -> None:
        tpl = self.repository.get_template(int(template_id))
        if not tpl:
            return
        uri, err = self.generate_template_preview(tpl.configuracion, format="thumbnail", quality="medium", page_number=1, sample_data=None)
        if err or not uri:
            return
        self.repository.update_template(int(template_id), updates={"preview_url": uri}, creada_por=None, cambios_descripcion=None)

    def get_template_versions(self, template_id: int) -> List[Dict[str, Any]]:
        versions = self.repository.get_template_versions(int(template_id))
        return [self._version_to_dict(v) for v in versions]

    def create_template_version(self, template_id: int, version: str, configuracion: Dict[str, Any], descripcion: Optional[str], creada_por: Optional[int]) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        if isinstance(configuracion, dict):
            configuracion = self._normalize_metadata_name(configuracion)
        v, err = self.repository.create_template_version(int(template_id), version=str(version), configuracion=configuracion, cambios_descripcion=descripcion, creada_por=creada_por)
        if not v:
            return None, err or "create_version_failed"
        return self._version_to_dict(v), None

    def restore_template_version(self, template_id: int, version: str) -> Tuple[bool, Optional[str]]:
        return self.repository.restore_template_version(int(template_id), version=str(version))

    def get_template_analytics(self, template_id: int, days: int = 30) -> Dict[str, Any]:
        return self.repository.get_template_analytics(int(template_id), days=int(days or 30))

    def get_analytics_dashboard(self, days: int = 30) -> Dict[str, Any]:
        days = int(days or 30)
        now = datetime.utcnow()
        inicio = now - timedelta(days=max(1, days))

        export_filter = [
            PlantillaRutina.configuracion.has_key("pages"),  # type: ignore[attr-defined]
            PlantillaRutina.configuracion.has_key("metadata"),  # type: ignore[attr-defined]
            PlantillaRutina.configuracion.has_key("layout"),  # type: ignore[attr-defined]
            PlantillaRutina.configuracion.has_key("variables"),  # type: ignore[attr-defined]
        ]

        total_templates = int(self.db.query(func.count(PlantillaRutina.id)).filter(*export_filter).scalar() or 0)
        active_templates = int(
            self.db.query(func.count(PlantillaRutina.id)).filter(*export_filter, PlantillaRutina.activa == True).scalar() or 0
        )

        total_events = int(
            self.db.query(func.count(PlantillaAnalitica.id))
            .filter(PlantillaAnalitica.fecha_evento >= inicio)
            .scalar()
            or 0
        )
        unique_users = int(
            self.db.query(func.count(distinct(PlantillaAnalitica.usuario_id)))
            .filter(PlantillaAnalitica.fecha_evento >= inicio, PlantillaAnalitica.usuario_id.isnot(None))
            .scalar()
            or 0
        )

        category_rows = (
            self.db.query(PlantillaRutina.categoria, func.count(PlantillaRutina.id))
            .filter(*export_filter)
            .group_by(PlantillaRutina.categoria)
            .order_by(func.count(PlantillaRutina.id).desc())
            .all()
        )
        category_analytics = [{"categoria": str(c or "general"), "count": int(n or 0)} for (c, n) in category_rows]

        popular_rows = (
            self.db.query(PlantillaRutina.id, PlantillaRutina.nombre, PlantillaRutina.categoria, func.count(PlantillaAnalitica.id).label("cnt"))
            .join(PlantillaAnalitica, PlantillaAnalitica.plantilla_id == PlantillaRutina.id)
            .filter(PlantillaAnalitica.fecha_evento >= inicio)
            .group_by(PlantillaRutina.id, PlantillaRutina.nombre, PlantillaRutina.categoria)
            .order_by(func.count(PlantillaAnalitica.id).desc())
            .limit(10)
            .all()
        )
        popular_templates = [
            {"template_id": int(tid), "nombre": str(nombre), "categoria": str(cat or "general"), "count": int(cnt or 0)}
            for (tid, nombre, cat, cnt) in popular_rows
        ]

        return {
            "overview": {
                "total_templates": total_templates,
                "active_templates": active_templates,
                "total_events": total_events,
                "unique_users": unique_users,
            },
            "popular_templates": popular_templates,
            "category_analytics": category_analytics,
        }

    def export_template(self, template_id: int) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        tpl = self.repository.get_template(int(template_id))
        if not tpl:
            return None, "Template not found"
        return {"template": self._template_to_dict(tpl)}, None

    def create_template_from_excel(
        self,
        excel_bytes: bytes,
        image_bytes: Optional[bytes],
        payload: Dict[str, Any],
        creada_por: Optional[int] = None,
    ) -> Tuple[Optional[Dict[str, Any]], Optional[str]]:
        nombre = str(payload.get("nombre") or "Plantilla Excel Importada").strip()
        if not nombre:
            nombre = "Plantilla Excel Importada"
        descripcion = str(payload.get("descripcion") or "").strip()
        if not descripcion:
            descripcion = f"Plantilla importada desde Excel ({nombre})"
        categoria = str(payload.get("categoria") or "general")
        tags = payload.get("tags")
        if tags is not None and not isinstance(tags, list):
            tags = None
        publica = bool(payload.get("publica") or False)
        activa = bool(payload.get("activa") if payload.get("activa") is not None else True)
        replace_defaults = bool(payload.get("replace_defaults") if payload.get("replace_defaults") is not None else True)
        dias_semana = self._detect_days_from_excel(excel_bytes)
        try:
            override_dias = payload.get("dias_semana")
            if override_dias is not None:
                dias_semana = int(override_dias)
        except Exception:
            pass
        if dias_semana is None or dias_semana <= 0:
            dias_semana = 3
        dias_semana = max(2, min(5, int(dias_semana)))
        orientation = self._infer_orientation_from_image(image_bytes)
        headers = self._extract_excel_headers(excel_bytes)
        weeks, week_columns = self._extract_weeks_and_columns(excel_bytes)
        configuracion = self._build_excel_equivalent_config(
            nombre=nombre,
            descripcion=descripcion,
            categoria=categoria,
            dias_semana=dias_semana,
            orientation=orientation,
            headers=headers,
            weeks=weeks,
            week_columns=week_columns,
        )
        configuracion = self._normalize_metadata_name(configuracion)
        tpl, err = self.repository.create_template(
            nombre=nombre,
            configuracion=configuracion,
            descripcion=descripcion,
            categoria=categoria,
            dias_semana=dias_semana,
            creada_por=creada_por,
            tags=[str(t) for t in tags] if isinstance(tags, list) else None,
            publica=publica,
            activa=activa,
        )
        if not tpl:
            return None, err or "create_failed"
        if replace_defaults:
            try:
                self._replace_default_excel_templates(configuracion, creada_por=creada_por)
            except Exception:
                pass
        return self._template_to_dict(tpl), None

    def _build_excel_equivalent_config(
        self,
        nombre: str,
        descripcion: str,
        categoria: str,
        dias_semana: int,
        orientation: Optional[str],
        headers: List[str],
        weeks: Optional[int],
        week_columns: Optional[List[str]],
    ) -> Dict[str, Any]:
        base_tags = ["excel", "imported", f"{dias_semana}_dias"]
        metadata_tags = list(dict.fromkeys(base_tags))
        layout_orientation = orientation if orientation in ("portrait", "landscape") else "portrait"
        total_weeks = int(weeks or 4)
        week_headers = week_columns if week_columns else ["Ser", "Rep", "Kg", "RIR"]
        config = {
            "metadata": {
                "name": nombre,
                "version": "1.0.0",
                "description": descripcion,
                "author": "import",
                "category": categoria,
                "difficulty": "beginner",
                "tags": metadata_tags,
                "estimated_duration": 45,
            },
            "layout": {
                "page_size": "A4",
                "orientation": layout_orientation,
                "margins": {"top": 20, "bottom": 20, "left": 20, "right": 20},
            },
            "pages": [
                {
                    "name": "Rutina",
                    "sections": [
                        {"type": "excel_header", "content": {"weeks": total_weeks}},
                        {"type": "spacing", "content": {"height": 8}},
                        {
                            "type": "exercise_table",
                            "content": {
                                "format": "excel_weekly",
                                "weeks": total_weeks,
                                "week_columns": week_headers,
                                "label": "EJERCICIOS",
                            },
                        },
                        {"type": "spacing", "content": {"height": 12}},
                        {"type": "qr_code", "content": {"size": 90}},
                    ],
                }
            ],
            "variables": {
                "gym_name": {"type": "string", "default": "Gimnasio", "required": False},
                "nombre_rutina": {"type": "string", "default": "Rutina", "required": False},
                "usuario_nombre": {"type": "string", "default": "Usuario", "required": False},
                "fecha": {"type": "string", "default": "", "required": False},
                "current_year": {"type": "string", "default": "", "required": False},
                "gym_logo_base64": {"type": "string", "default": "", "required": False},
                "total_weeks": {"type": "number", "default": total_weeks, "required": False},
            },
            "qr_code": {"enabled": True, "position": "inline", "data_source": "routine_uuid"},
            "styling": {
                "fonts": {
                    "title": {"family": "Helvetica-Bold", "size": 18, "color": "#000000"},
                    "body": {"family": "Helvetica", "size": 10, "color": "#111827"},
                },
                "colors": {"primary": "#111827", "accent": "#3B82F6"},
            },
            "excel_equivalent": f"Import_{dias_semana}_dias",
            "dias_semana": dias_semana,
        }
        if headers:
            config["excel_headers"] = headers
        return config

    def _infer_orientation_from_image(self, image_bytes: Optional[bytes]) -> Optional[str]:
        if not image_bytes:
            return None
        try:
            with Image.open(BytesIO(image_bytes)) as img:
                width, height = img.size
            if width > height:
                return "landscape"
            return "portrait"
        except Exception:
            return None

    def _extract_excel_headers(self, excel_bytes: bytes) -> List[str]:
        try:
            from openpyxl import load_workbook
        except Exception:
            return []
        try:
            wb = load_workbook(BytesIO(excel_bytes), data_only=True)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, max_row=25, values_only=True):
                if not row:
                    continue
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if len(values) >= 2:
                    headers = []
                    for v in row:
                        if v is None:
                            headers.append("")
                        else:
                            headers.append(str(v).strip())
                    cleaned = [h for h in headers if h]
                    return cleaned
            return []
        except Exception:
            return []

    def _extract_weeks_and_columns(self, excel_bytes: bytes) -> Tuple[Optional[int], Optional[List[str]]]:
        try:
            from openpyxl import load_workbook
        except Exception:
            return None, None
        try:
            wb = load_workbook(BytesIO(excel_bytes), data_only=True)
            ws = wb.active
            week_row_idx = None
            week_cols: List[int] = []
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=30, values_only=True), start=1):
                if not row:
                    continue
                found = False
                for col_idx, cell in enumerate(row, start=1):
                    if cell is None:
                        continue
                    raw = str(cell).strip().lower()
                    if raw.startswith("semana"):
                        found = True
                        week_cols.append(col_idx)
                if found:
                    week_row_idx = idx
                    break
            if not week_cols:
                return None, None
            week_cols = sorted(set(week_cols))
            weeks_count = len(week_cols)
            subheaders: List[str] = []
            if week_row_idx and week_row_idx + 1 <= ws.max_row:
                row = [ws.cell(week_row_idx + 1, c).value for c in range(1, ws.max_column + 1)]
                start = week_cols[0]
                if len(week_cols) > 1:
                    end = week_cols[1] - 1
                else:
                    end = min(ws.max_column, start + 10)
                for c in range(start, end + 1):
                    val = row[c - 1] if c - 1 < len(row) else None
                    if val is None:
                        continue
                    raw = str(val).strip()
                    if raw:
                        subheaders.append(raw)
            if not subheaders:
                subheaders = ["Ser", "Rep", "Kg", "RIR"]
            return weeks_count, subheaders
        except Exception:
            return None, None

    def _replace_default_excel_templates(self, configuracion: Dict[str, Any], creada_por: Optional[int]) -> None:
        names = [
            "Plantilla Excel 2 días",
            "Plantilla Excel 3 días",
            "Plantilla Excel 4 días",
            "Plantilla Excel 5 días",
        ]
        for name in names:
            try:
                tpl = self.db.query(PlantillaRutina).filter(PlantillaRutina.nombre == name).first()
                if not tpl:
                    continue
                descripcion = str(tpl.descripcion or "").strip()
                if "dinámica" not in descripcion.lower():
                    descripcion = (descripcion + " " if descripcion else "") + "Reemplazada por plantilla dinámica"
                updates = {
                    "configuracion": configuracion,
                    "descripcion": descripcion,
                    "publica": False,
                    "activa": True,
                }
                self.repository.update_template(
                    int(tpl.id),
                    updates=updates,
                    creada_por=creada_por,
                    cambios_descripcion="Reemplazo por plantilla dinámica",
                )
            except Exception:
                continue

    def _detect_days_from_excel(self, excel_bytes: bytes) -> Optional[int]:
        try:
            from openpyxl import load_workbook
        except Exception:
            return None
        try:
            wb = load_workbook(BytesIO(excel_bytes), data_only=True)
            ws = wb.active
            header_row_idx = None
            header_values: List[str] = []
            for idx, row in enumerate(ws.iter_rows(min_row=1, max_row=25, values_only=True), start=1):
                if not row:
                    continue
                values = [str(v).strip() for v in row if v is not None and str(v).strip()]
                if len(values) >= 2:
                    header_row_idx = idx
                    header_values = [str(v).strip() if v is not None else "" for v in row]
                    break
            day_col = None
            if header_values:
                for i, h in enumerate(header_values):
                    if not h:
                        continue
                    hl = h.strip().lower()
                    if "dia" in hl or "día" in hl or "day" in hl:
                        day_col = i
                        break
            days_found: List[int] = []
            if day_col is not None and header_row_idx is not None:
                for row in ws.iter_rows(min_row=header_row_idx + 1, max_row=header_row_idx + 300, values_only=True):
                    if not row:
                        continue
                    raw = row[day_col] if day_col < len(row) else None
                    if raw is None:
                        continue
                    try:
                        day_val = int(str(raw).strip())
                    except Exception:
                        continue
                    if day_val > 0:
                        days_found.append(day_val)
            if days_found:
                return len(sorted(set(days_found)))
            day_names = {
                "lunes": 1,
                "martes": 2,
                "miercoles": 3,
                "miércoles": 3,
                "jueves": 4,
                "viernes": 5,
                "sabado": 6,
                "sábado": 6,
                "domingo": 7,
            }
            day_matches: List[int] = []
            for row in ws.iter_rows(min_row=1, max_row=200, values_only=True):
                for cell in row:
                    if cell is None:
                        continue
                    raw = str(cell).strip().lower()
                    if not raw:
                        continue
                    for name, num in day_names.items():
                        if raw.startswith(name):
                            day_matches.append(num)
                    if raw.startswith("dia ") or raw.startswith("día "):
                        parts = raw.replace("día", "dia").split()
                        if len(parts) >= 2:
                            try:
                                day_matches.append(int(parts[1]))
                            except Exception:
                                pass
            if day_matches:
                return len(sorted(set(day_matches)))
            return None
        except Exception:
            return None

    def _template_to_dict(self, tpl) -> Dict[str, Any]:
        return {
            "id": int(tpl.id),
            "nombre": str(tpl.nombre),
            "descripcion": tpl.descripcion,
            "configuracion": tpl.configuracion,
            "categoria": str(tpl.categoria or "general"),
            "dias_semana": tpl.dias_semana,
            "activa": bool(tpl.activa),
            "publica": bool(tpl.publica),
            "creada_por": tpl.creada_por,
            "fecha_creacion": tpl.fecha_creacion.isoformat() if getattr(tpl, "fecha_creacion", None) else None,
            "fecha_actualizacion": tpl.fecha_actualizacion.isoformat() if getattr(tpl, "fecha_actualizacion", None) else None,
            "version_actual": str(tpl.version_actual or "1.0.0"),
            "tags": list(tpl.tags or []),
            "preview_url": tpl.preview_url,
            "uso_count": int(tpl.uso_count or 0),
            "rating_promedio": float(tpl.rating_promedio) if tpl.rating_promedio is not None else None,
            "rating_count": int(tpl.rating_count or 0),
        }

    def _version_to_dict(self, v) -> Dict[str, Any]:
        return {
            "id": int(v.id),
            "plantilla_id": int(v.plantilla_id),
            "version": str(v.version),
            "configuracion": v.configuracion,
            "cambios_descripcion": v.cambios_descripcion,
            "creada_por": v.creada_por,
            "fecha_creacion": v.fecha_creacion.isoformat() if getattr(v, "fecha_creacion", None) else None,
            "es_actual": bool(v.es_actual),
        }
