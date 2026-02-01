from __future__ import annotations

import csv
import io
import json
import threading
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Tuple

from fastapi import APIRouter, BackgroundTasks, Depends, File, HTTPException, Request, UploadFile
from fastapi.responses import JSONResponse, StreamingResponse
from openpyxl import Workbook, load_workbook
from sqlalchemy import text
from sqlalchemy.orm import Session

from src.database import get_tenant_session_factory
from src.dependencies import get_db_session, get_claims, require_feature, require_owner
from src.services.user_service import UserService

router = APIRouter(
    dependencies=[
        Depends(require_owner),
        Depends(require_feature("bulk_actions")),
        Depends(require_feature("bulk_actions:usuarios_import")),
    ]
)


KIND_USUARIOS_IMPORT = "usuarios_import"
SUPPORTED_KINDS = {KIND_USUARIOS_IMPORT}


def _normalize_header(h: Any) -> str:
    return str(h or "").strip().lower().replace(" ", "_")


def _boolish(v: Any) -> Optional[bool]:
    if v is None:
        return None
    if isinstance(v, bool):
        return bool(v)
    s = str(v).strip().lower()
    if s in ("1", "true", "t", "si", "sí", "s", "yes", "y"):
        return True
    if s in ("0", "false", "f", "no", "n"):
        return False
    return None


def _digits(v: Any) -> str:
    s = str(v or "").strip()
    return "".join(ch for ch in s if ch.isdigit())


def _iter_csv_rows(content: bytes):
    text_content = content.decode("utf-8-sig", errors="replace")
    reader = csv.reader(io.StringIO(text_content))
    it = iter(reader)
    try:
        header_row = next(it)
    except StopIteration:
        return ([], iter(()))
    headers = [str(x or "").strip() for x in list(header_row)]
    return (headers, it)


def _iter_xlsx_rows(content: bytes):
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows_iter)
    except StopIteration:
        return ([], iter(()))
    headers = [str(x or "").strip() for x in list(header_row)]
    return (headers, rows_iter)


def _resolve_tipo_cuota_map(db: Session) -> Dict[str, int]:
    rows = db.execute(text("SELECT id, nombre FROM tipos_cuota")).fetchall()
    m: Dict[str, int] = {}
    for r in rows:
        try:
            m[str(r[1] or "").strip().lower()] = int(r[0])
        except Exception:
            continue
    return m


def _validate_usuario_row(
    raw: Dict[str, Any], *, tipo_cuota_map: Dict[str, int]
) -> Tuple[Dict[str, Any], List[str], List[str]]:
    errors: List[str] = []
    warnings: List[str] = []

    nombre = str(raw.get("nombre") or "").strip()
    dni = str(raw.get("dni") or "").strip()
    telefono = str(raw.get("telefono") or "").strip()

    dni_norm = _digits(dni) if dni else ""
    tel_norm = _digits(telefono) if telefono else ""

    if not nombre:
        errors.append("nombre es requerido")
    if not dni_norm and not tel_norm:
        errors.append("dni o telefono es requerido")

    tipo_cuota_id: Optional[int] = None
    tc_id_raw = raw.get("tipo_cuota_id")
    tc_name_raw = raw.get("tipo_cuota")

    if tc_id_raw not in (None, ""):
        try:
            tipo_cuota_id = int(str(tc_id_raw).strip())
        except Exception:
            errors.append("tipo_cuota_id inválido")
    elif tc_name_raw not in (None, ""):
        key = str(tc_name_raw).strip().lower()
        tipo_cuota_id = tipo_cuota_map.get(key)
        if tipo_cuota_id is None:
            errors.append("tipo_cuota no existe")

    activo = _boolish(raw.get("activo"))
    if activo is None:
        activo = True

    notas = raw.get("notas")
    if notas is not None:
        notas = str(notas).strip()
        if notas == "":
            notas = None

    out = {
        "nombre": nombre.upper(),
        "dni": dni_norm,
        "telefono": tel_norm,
        "tipo_cuota_id": tipo_cuota_id,
        "activo": bool(activo),
        "notas": notas,
    }
    return (out, errors, warnings)


def _insert_job(
    db: Session,
    *,
    kind: str,
    request: Request,
    filename: Optional[str],
    mime: Optional[str],
    size_bytes: Optional[int],
) -> int:
    claims = get_claims(request)
    role = str(claims.get("role") or "").strip().lower()
    created_by_user_id = claims.get("user_id")
    try:
        created_by_user_id_int = int(created_by_user_id) if created_by_user_id else None
    except Exception:
        created_by_user_id_int = None
    job_id = db.execute(
        text(
            """
            INSERT INTO bulk_jobs(kind, status, created_by_user_id, created_by_role, source_filename, source_mime, source_size_bytes)
            VALUES (:kind, 'draft', :uid, :role, :fn, :mime, :sz)
            RETURNING id
            """
        ),
        {
            "kind": kind,
            "uid": created_by_user_id_int,
            "role": role,
            "fn": filename,
            "mime": mime,
            "sz": int(size_bytes) if size_bytes is not None else None,
        },
    ).scalar()
    db.commit()
    return int(job_id)


def _save_rows(
    db: Session, job_id: int, rows: List[Dict[str, Any]], validations: List[Tuple[List[str], List[str]]]
) -> None:
    params: List[Dict[str, Any]] = []
    for idx, data in enumerate(rows):
        errs, warns = validations[idx]
        params.append(
            {
                "job_id": int(job_id),
                "row_index": int(idx),
                "data": json.dumps(data or {}, ensure_ascii=False),
                "errors": json.dumps(errs or [], ensure_ascii=False),
                "warnings": json.dumps(warns or [], ensure_ascii=False),
                "is_valid": bool(len(errs) == 0),
            }
        )
    db.execute(
        text(
            """
            INSERT INTO bulk_job_rows(job_id, row_index, data, errors, warnings, is_valid)
            VALUES (:job_id, :row_index, CAST(:data AS JSONB), CAST(:errors AS JSONB), CAST(:warnings AS JSONB), :is_valid)
            """
        ),
        params,
    )
    valid = sum(1 for e, _w in validations if len(e) == 0)
    invalid = len(rows) - valid
    db.execute(
        text(
            """
            UPDATE bulk_jobs
            SET rows_total = :total, rows_valid = :valid, rows_invalid = :invalid, updated_at = NOW()
            WHERE id = :id
            """
        ),
        {"id": int(job_id), "total": int(len(rows)), "valid": int(valid), "invalid": int(invalid)},
    )
    db.commit()


def _job_summary(db: Session, job_id: int) -> Dict[str, Any]:
    job = db.execute(
        text(
            """
            SELECT id, kind, status, created_at, updated_at, rows_total, rows_valid, rows_invalid, applied_count, error_count, failure_reason
            FROM bulk_jobs
            WHERE id = :id
            """
        ),
        {"id": int(job_id)},
    ).mappings().first()
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    return dict(job)


def _run_job_background(*, tenant: str, job_id: int) -> None:
    SessionFactory = get_tenant_session_factory(tenant)
    db = SessionFactory()
    try:
        kind = db.execute(text("SELECT kind FROM bulk_jobs WHERE id = :id"), {"id": int(job_id)}).scalar()
        if str(kind) != KIND_USUARIOS_IMPORT:
            db.execute(
                text("UPDATE bulk_jobs SET status = 'failed', failure_reason = :r, updated_at = NOW() WHERE id = :id"),
                {"id": int(job_id), "r": "Tipo de job no soportado"},
            )
            db.commit()
            return

        got_lock = db.execute(text("SELECT pg_try_advisory_lock(hashtext('bulk_actions_global'))")).scalar()
        if not bool(got_lock):
            db.execute(
                text("UPDATE bulk_jobs SET status = 'failed', failure_reason = :r, updated_at = NOW() WHERE id = :id"),
                {"id": int(job_id), "r": "Otro job masivo está en ejecución"},
            )
            db.commit()
            return

        db.execute(text("UPDATE bulk_jobs SET status = 'running', updated_at = NOW() WHERE id = :id"), {"id": int(job_id)})
        db.commit()

        svc = UserService(db)
        rows = db.execute(
            text(
                """
                SELECT row_index, data
                FROM bulk_job_rows
                WHERE job_id = :id AND is_valid = true
                ORDER BY row_index ASC
                """
            ),
            {"id": int(job_id)},
        ).fetchall()

        applied = 0
        failed = 0
        for row_index, data in rows:
            try:
                if (applied + failed) % 50 == 0:
                    st = db.execute(
                        text("SELECT status FROM bulk_jobs WHERE id = :id"),
                        {"id": int(job_id)},
                    ).scalar()
                    if str(st or "") == "cancelled":
                        db.execute(
                            text(
                                "UPDATE bulk_jobs SET status = 'cancelled', updated_at = NOW() WHERE id = :id"
                            ),
                            {"id": int(job_id)},
                        )
                        db.commit()
                        return
            except Exception:
                pass
            try:
                dni = str((data or {}).get("dni") or "").strip()
                tel = str((data or {}).get("telefono") or "").strip()

                existing = None
                if dni:
                    existing = svc.get_user_by_dni(dni)
                if (existing is None) and tel:
                    try:
                        existing = svc.user_repo.obtener_usuario_por_telefono(tel)
                    except Exception:
                        existing = None

                if existing is None:
                    payload = dict(data or {})
                    payload["rol"] = "socio"
                    payload["fecha_registro"] = datetime.now(timezone.utc).replace(tzinfo=None)
                    svc.create_user(payload, is_owner=True)
                else:
                    uid = int(getattr(existing, "id"))
                    payload = dict(data or {})
                    payload.pop("dni", None)
                    svc.update_user(uid, payload, modifier_id=None, is_owner=True)

                applied += 1
                db.execute(
                    text(
                        """
                        UPDATE bulk_job_rows
                        SET applied = true, applied_at = NOW(), result = CAST(:res AS JSONB)
                        WHERE job_id = :jid AND row_index = :idx
                        """
                    ),
                    {"jid": int(job_id), "idx": int(row_index), "res": json.dumps({"ok": True}, ensure_ascii=False)},
                )
            except Exception as e:
                failed += 1
                db.execute(
                    text(
                        """
                        UPDATE bulk_job_rows
                        SET applied = false, result = CAST(:res AS JSONB)
                        WHERE job_id = :jid AND row_index = :idx
                        """
                    ),
                    {"jid": int(job_id), "idx": int(row_index), "res": json.dumps({"ok": False, "error": str(e)}, ensure_ascii=False)},
                )

            if (applied + failed) % 50 == 0:
                db.execute(
                    text(
                        "UPDATE bulk_jobs SET applied_count = :a, error_count = :e, updated_at = NOW() WHERE id = :id"
                    ),
                    {"id": int(job_id), "a": int(applied), "e": int(failed)},
                )
                db.commit()

        db.execute(
            text("UPDATE bulk_jobs SET applied_count = :a, error_count = :e, status = 'completed', updated_at = NOW() WHERE id = :id"),
            {"id": int(job_id), "a": int(applied), "e": int(failed)},
        )
        db.commit()
    except Exception as e:
        try:
            db.execute(
                text("UPDATE bulk_jobs SET status = 'failed', failure_reason = :r, updated_at = NOW() WHERE id = :id"),
                {"id": int(job_id), "r": str(e)},
            )
            db.commit()
        except Exception:
            pass
    finally:
        try:
            db.execute(text("SELECT pg_advisory_unlock(hashtext('bulk_actions_global'))"))
            db.commit()
        except Exception:
            pass
        try:
            db.close()
        except Exception:
            pass


@router.get("/api/bulk/templates/{kind}.{fmt}")
async def api_bulk_template(kind: str, fmt: str, db: Session = Depends(get_db_session)):
    k = str(kind or "").strip().lower()
    f = str(fmt or "").strip().lower()
    if k not in SUPPORTED_KINDS:
        raise HTTPException(status_code=404, detail="Tipo no soportado")
    if f not in ("csv", "xlsx"):
        raise HTTPException(status_code=400, detail="Formato inválido")

    headers = ["nombre", "dni", "telefono", "tipo_cuota", "activo", "notas"]
    sample = ["JUAN PEREZ", "12345678", "1122334455", "", "true", ""]

    if f == "csv":
        out = io.StringIO()
        w = csv.writer(out)
        w.writerow(headers)
        w.writerow(sample)
        data = out.getvalue().encode("utf-8")
        return StreamingResponse(
            io.BytesIO(data),
            media_type="text/csv",
            headers={"Content-Disposition": f'attachment; filename="{k}.csv"'},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = "template"
    ws.append(headers)
    ws.append(sample)
    try:
        ws2 = wb.create_sheet("tipos_cuota")
        ws2.append(["nombre"])
        items = db.execute(text("SELECT nombre FROM tipos_cuota WHERE activo = true ORDER BY nombre ASC")).fetchall()
        for it in items or []:
            ws2.append([str(it[0] or "").strip()])
    except Exception:
        pass
    bio = io.BytesIO()
    wb.save(bio)
    bio.seek(0)
    return StreamingResponse(
        bio,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{k}.xlsx"'},
    )


@router.post("/api/bulk/jobs/preview")
async def api_bulk_preview(
    request: Request,
    kind: str,
    file: UploadFile = File(...),
    db: Session = Depends(get_db_session),
):
    k = str(kind or "").strip().lower()
    if k not in SUPPORTED_KINDS:
        raise HTTPException(status_code=404, detail="Tipo no soportado")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Archivo vacío")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=400, detail="Archivo demasiado grande (max 10MB)")

    name = str(file.filename or "").lower()
    if name.endswith(".xlsx"):
        headers, rows_iter = _iter_xlsx_rows(content)
    else:
        headers, rows_iter = _iter_csv_rows(content)

    if not headers:
        raise HTTPException(status_code=400, detail="No se detectaron columnas")

    alias = {
        "documento": "dni",
        "doc": "dni",
        "tel": "telefono",
        "celular": "telefono",
        "plan": "tipo_cuota",
        "tipo_cuota_nombre": "tipo_cuota",
    }
    cols = [_normalize_header(h) for h in headers]
    cols = [alias.get(c, c) for c in cols]

    tipo_cuota_map = _resolve_tipo_cuota_map(db)

    job_id = _insert_job(
        db,
        kind=k,
        request=request,
        filename=file.filename,
        mime=file.content_type,
        size_bytes=len(content),
    )

    max_rows = 20000
    batch_size = 250

    total = 0
    valid = 0
    invalid = 0
    preview_rows: List[Dict[str, Any]] = []

    batch_params: List[Dict[str, Any]] = []
    row_index = 0
    for r in rows_iter:
        if r is None:
            continue
        try:
            row_vals = list(r)
        except Exception:
            row_vals = []
        raw = {cols[i]: (row_vals[i] if i < len(row_vals) else None) for i in range(len(cols))}
        if all((raw.get(c) in (None, "", 0) for c in cols)):
            continue

        if k == KIND_USUARIOS_IMPORT:
            out, errs, warns = _validate_usuario_row(raw, tipo_cuota_map=tipo_cuota_map)
        else:
            out, errs, warns = ({}, ["tipo no soportado"], [])

        is_valid = bool(len(errs) == 0)
        total += 1
        if is_valid:
            valid += 1
        else:
            invalid += 1

        if len(preview_rows) < 200:
            preview_rows.append({"row_index": int(row_index), "data": out, "errors": errs, "warnings": warns})

        batch_params.append(
            {
                "job_id": int(job_id),
                "row_index": int(row_index),
                "data": json.dumps(out or {}, ensure_ascii=False),
                "errors": json.dumps(errs or [], ensure_ascii=False),
                "warnings": json.dumps(warns or [], ensure_ascii=False),
                "is_valid": bool(is_valid),
            }
        )

        row_index += 1
        if len(batch_params) >= batch_size:
            db.execute(
                text(
                    """
                    INSERT INTO bulk_job_rows(job_id, row_index, data, errors, warnings, is_valid)
                    VALUES (:job_id, :row_index, CAST(:data AS JSONB), CAST(:errors AS JSONB), CAST(:warnings AS JSONB), :is_valid)
                    """
                ),
                batch_params,
            )
            db.commit()
            batch_params = []

        if total >= max_rows:
            break

    if batch_params:
        db.execute(
            text(
                """
                INSERT INTO bulk_job_rows(job_id, row_index, data, errors, warnings, is_valid)
                VALUES (:job_id, :row_index, CAST(:data AS JSONB), CAST(:errors AS JSONB), CAST(:warnings AS JSONB), :is_valid)
                """
            ),
            batch_params,
        )
        db.commit()

    db.execute(
        text(
            """
            UPDATE bulk_jobs
            SET rows_total = :total, rows_valid = :valid, rows_invalid = :invalid, updated_at = NOW()
            WHERE id = :id
            """
        ),
        {"id": int(job_id), "total": int(total), "valid": int(valid), "invalid": int(invalid)},
    )
    db.commit()

    return {
        "ok": True,
        "job": _job_summary(db, job_id),
        "preview": {
            "columns": list(preview_rows[0]["data"].keys()) if preview_rows else [],
            "rows": preview_rows,
        },
    }


@router.get("/api/bulk/jobs/{job_id}")
async def api_bulk_job_get(job_id: int, db: Session = Depends(get_db_session)):
    job = _job_summary(db, job_id)
    recent_rows = db.execute(
        text(
            """
            SELECT row_index, data, errors, warnings, is_valid, applied, result
            FROM bulk_job_rows
            WHERE job_id = :id
            ORDER BY row_index ASC
            LIMIT 200
            """
        ),
        {"id": int(job_id)},
    ).mappings().all()
    return {"ok": True, "job": job, "rows": [dict(r) for r in recent_rows]}


@router.get("/api/bulk/jobs")
async def api_bulk_jobs_list(
    db: Session = Depends(get_db_session),
    kind: str = "",
    status: str = "",
    page: int = 1,
    limit: int = 20,
):
    page_i = max(1, int(page or 1))
    limit_i = max(1, min(int(limit or 20), 100))
    offset_i = (page_i - 1) * limit_i
    where: List[str] = []
    params: Dict[str, Any] = {"limit": int(limit_i), "offset": int(offset_i)}
    if str(kind or "").strip():
        where.append("kind = :kind")
        params["kind"] = str(kind).strip().lower()
    if str(status or "").strip():
        where.append("status = :status")
        params["status"] = str(status).strip().lower()
    where_sql = " WHERE " + " AND ".join(where) if where else ""
    total = db.execute(text(f"SELECT COUNT(*) FROM bulk_jobs{where_sql}"), params).scalar() or 0
    rows = db.execute(
        text(
            f"""
            SELECT id, kind, status, created_at, updated_at, rows_total, rows_valid, rows_invalid, applied_count, error_count, failure_reason
            FROM bulk_jobs
            {where_sql}
            ORDER BY id DESC
            LIMIT :limit OFFSET :offset
            """
        ),
        params,
    ).mappings().all()
    return {"ok": True, "items": [dict(r) for r in rows], "total": int(total), "limit": int(limit_i), "offset": int(offset_i)}


@router.get("/api/bulk/jobs/{job_id}/errors.csv")
async def api_bulk_job_errors_csv(job_id: int, db: Session = Depends(get_db_session)):
    _ = _job_summary(db, job_id)
    rows = db.execute(
        text(
            """
            SELECT row_index, data, errors, warnings
            FROM bulk_job_rows
            WHERE job_id = :id AND is_valid = false
            ORDER BY row_index ASC
            """
        ),
        {"id": int(job_id)},
    ).mappings().all()
    out = io.StringIO()
    w = csv.writer(out)
    w.writerow(["row", "errors", "warnings", "data"])
    for r in rows:
        w.writerow(
            [
                int(r.get("row_index") or 0) + 1,
                json.dumps(r.get("errors") or [], ensure_ascii=False),
                json.dumps(r.get("warnings") or [], ensure_ascii=False),
                json.dumps(r.get("data") or {}, ensure_ascii=False),
            ]
        )
    data = out.getvalue().encode("utf-8")
    return StreamingResponse(
        io.BytesIO(data),
        media_type="text/csv",
        headers={"Content-Disposition": f'attachment; filename="bulk_job_{int(job_id)}_errors.csv"'},
    )


@router.put("/api/bulk/jobs/{job_id}/rows/{row_index}")
async def api_bulk_job_row_update(
    job_id: int,
    row_index: int,
    request: Request,
    db: Session = Depends(get_db_session),
):
    job = db.execute(
        text("SELECT kind, status FROM bulk_jobs WHERE id = :id"),
        {"id": int(job_id)},
    ).fetchone()
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    kind, status = job[0], str(job[1] or "")
    if status not in ("draft", "validated"):
        raise HTTPException(status_code=409, detail="Job no editable")

    try:
        payload = await request.json()
    except Exception:
        payload = {}
    data_in = payload.get("data") if isinstance(payload, dict) else None
    if not isinstance(data_in, dict):
        raise HTTPException(status_code=400, detail="data inválida")

    tipo_cuota_map = _resolve_tipo_cuota_map(db)
    if str(kind) == KIND_USUARIOS_IMPORT:
        out, errs, warns = _validate_usuario_row(data_in, tipo_cuota_map=tipo_cuota_map)
    else:
        out, errs, warns = ({}, ["tipo no soportado"], [])

    db.execute(
        text(
            """
            UPDATE bulk_job_rows
            SET data = CAST(:data AS JSONB),
                errors = CAST(:errors AS JSONB),
                warnings = CAST(:warnings AS JSONB),
                is_valid = :is_valid
            WHERE job_id = :jid AND row_index = :idx
            """
        ),
        {
            "jid": int(job_id),
            "idx": int(row_index),
            "data": json.dumps(out or {}, ensure_ascii=False),
            "errors": json.dumps(errs or [], ensure_ascii=False),
            "warnings": json.dumps(warns or [], ensure_ascii=False),
            "is_valid": bool(len(errs) == 0),
        },
    )

    counts = db.execute(
        text(
            """
            SELECT
                COUNT(*) AS total,
                SUM(CASE WHEN is_valid THEN 1 ELSE 0 END) AS valid,
                SUM(CASE WHEN NOT is_valid THEN 1 ELSE 0 END) AS invalid
            FROM bulk_job_rows
            WHERE job_id = :id
            """
        ),
        {"id": int(job_id)},
    ).mappings().first()
    db.execute(
        text(
            """
            UPDATE bulk_jobs
            SET rows_total = :total, rows_valid = :valid, rows_invalid = :invalid, updated_at = NOW()
            WHERE id = :id
            """
        ),
        {"id": int(job_id), "total": int(counts["total"] or 0), "valid": int(counts["valid"] or 0), "invalid": int(counts["invalid"] or 0)},
    )
    db.commit()
    return {"ok": True, "row_index": int(row_index), "data": out, "errors": errs, "warnings": warns, "job": _job_summary(db, job_id)}


@router.post("/api/bulk/jobs/{job_id}/confirm")
async def api_bulk_job_confirm(job_id: int, request: Request, db: Session = Depends(get_db_session)):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    token = str((payload or {}).get("confirmation") or "").strip().upper()
    if token != "CONFIRMAR":
        raise HTTPException(status_code=400, detail="Confirmación inválida")

    job = db.execute(text("SELECT status FROM bulk_jobs WHERE id = :id"), {"id": int(job_id)}).scalar()
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    if str(job) not in ("draft", "validated"):
        raise HTTPException(status_code=409, detail="Job no confirmable")

    invalid = db.execute(
        text("SELECT COUNT(*) FROM bulk_job_rows WHERE job_id = :id AND is_valid = false"),
        {"id": int(job_id)},
    ).scalar()
    if int(invalid or 0) > 0:
        raise HTTPException(status_code=400, detail="Hay filas inválidas, corregir antes de confirmar")

    db.execute(
        text("UPDATE bulk_jobs SET status = 'confirmed', updated_at = NOW() WHERE id = :id"),
        {"id": int(job_id)},
    )
    db.commit()
    return {"ok": True, "job": _job_summary(db, job_id)}


@router.post("/api/bulk/jobs/{job_id}/run")
async def api_bulk_job_run(
    job_id: int,
    request: Request,
    background: BackgroundTasks,
    db: Session = Depends(get_db_session),
):
    try:
        payload = await request.json()
    except Exception:
        payload = {}
    token = str((payload or {}).get("confirmation") or "").strip().upper()
    expected = payload.get("expected_rows")

    if token != "EJECUTAR":
        raise HTTPException(status_code=400, detail="Confirmación inválida")

    job = db.execute(text("SELECT status, rows_total FROM bulk_jobs WHERE id = :id"), {"id": int(job_id)}).fetchone()
    if not job:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    status, rows_total = str(job[0] or ""), int(job[1] or 0)
    if status != "confirmed":
        raise HTTPException(status_code=409, detail="Job debe estar confirmado")
    if expected is None or int(expected) != rows_total:
        raise HTTPException(status_code=400, detail="expected_rows inválido")

    claims = get_claims(request)
    tenant = str(claims.get("tenant") or "").strip()
    if not tenant:
        raise HTTPException(status_code=400, detail="Tenant inválido")

    t = threading.Thread(target=_run_job_background, kwargs={"tenant": tenant, "job_id": int(job_id)}, daemon=True)
    t.start()

    return {"ok": True, "job": _job_summary(db, job_id)}


@router.post("/api/bulk/jobs/{job_id}/cancel")
async def api_bulk_job_cancel(job_id: int, db: Session = Depends(get_db_session)):
    st = db.execute(text("SELECT status FROM bulk_jobs WHERE id = :id"), {"id": int(job_id)}).scalar()
    if not st:
        raise HTTPException(status_code=404, detail="Job no encontrado")
    s = str(st or "")
    if s in ("completed", "failed"):
        raise HTTPException(status_code=409, detail="Job no cancelable")
    db.execute(text("UPDATE bulk_jobs SET status = 'cancelled', updated_at = NOW() WHERE id = :id"), {"id": int(job_id)})
    db.commit()
    return {"ok": True, "job": _job_summary(db, job_id)}
