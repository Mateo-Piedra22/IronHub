"""
Excel Template Migration Tool
Converts existing Excel-based routine templates to the new dynamic template system
"""

import json
import logging
from typing import Dict, List, Any, Optional, Tuple
from dataclasses import dataclass, asdict
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

@dataclass
class ExcelTemplateConfig:
    """Configuration extracted from Excel template"""
    name: str
    description: str
    category: str = "general"
    days_per_week: Optional[int] = None
    layout_config: Dict[str, Any] = None
    styling_config: Dict[str, Any] = None
    exercise_config: Dict[str, Any] = None

@dataclass
class ExerciseMapping:
    """Mapping from Excel exercise to template exercise"""
    name: str
    sets: int
    reps: str
    rest: Optional[str] = None
    notes: Optional[str] = None
    muscle_group: Optional[str] = None
    difficulty: Optional[str] = None

class ExcelTemplateMigrator:
    """Migrates Excel templates to dynamic template system"""
    
    def __init__(self):
        self.excel_templates_dir = Path("excel_templates")
        self.output_dir = Path("migrated_templates")
        self.migration_log = []
        
        # Create output directory
        self.output_dir.mkdir(exist_ok=True)
        
        # Default template configuration
        self.default_config = {
            "version": "1.0.0",
            "metadata": {
                "migrated_from": "excel",
                "migration_date": datetime.now().isoformat(),
                "author": "System Migration"
            },
            "layout": {
                "page_size": "A4",
                "orientation": "portrait",
                "margins": {
                    "top": 20,
                    "right": 20,
                    "bottom": 20,
                    "left": 20
                }
            },
            "sections": [],
            "variables": {
                "gym_name": {
                    "type": "string",
                    "default": "Gym",
                    "description": "Nombre del gimnasio"
                },
                "client_name": {
                    "type": "string",
                    "default": "Cliente",
                    "description": "Nombre del cliente"
                },
                "trainer_name": {
                    "type": "string",
                    "default": "Entrenador",
                    "description": "Nombre del entrenador"
                }
            },
            "styling": {
                "primary_color": "#000000",
                "secondary_color": "#666666",
                "font_family": "Arial",
                "font_size": 12
            }
        }

    def scan_excel_templates(self) -> List[Path]:
        """Scan directory for Excel template files"""
        excel_files = []
        
        if not self.excel_templates_dir.exists():
            logger.warning(f"Excel templates directory not found: {self.excel_templates_dir}")
            return excel_files
            
        for ext in ['*.xlsx', '*.xls']:
            excel_files.extend(self.excel_templates_dir.glob(ext))
            
        logger.info(f"Found {len(excel_files)} Excel template files")
        return excel_files

    def analyze_excel_structure(self, file_path: Path) -> Dict[str, Any]:
        """Analyze Excel file structure to understand template format"""
        try:
            wb = load_workbook(file_path, read_only=True)
            structure = {
                "file_name": file_path.name,
                "sheets": [],
                "has_macros": False,
                "has_images": False,
                "total_rows": 0,
                "total_cols": 0
            }
            
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                sheet_info = {
                    "name": sheet_name,
                    "max_row": ws.max_row,
                    "max_col": ws.max_column,
                    "has_data": False,
                    "data_ranges": [],
                    "merged_cells": len(ws.merged_cells.ranges),
                    "has_formulas": False
                }
                
                # Analyze data ranges
                data_start = None
                data_end = None
                
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.value is not None:
                            sheet_info["has_data"] = True
                            if data_start is None:
                                data_start = (cell.row, cell.column)
                            data_end = (cell.row, cell.column)
                            
                            if isinstance(cell.value, str) and cell.value.startswith('='):
                                sheet_info["has_formulas"] = True
                
                if data_start and data_end:
                    sheet_info["data_ranges"].append({
                        "start": data_start,
                        "end": data_end
                    })
                
                structure["sheets"].append(sheet_info)
                structure["total_rows"] = max(structure["total_rows"], sheet_info["max_row"])
                structure["total_cols"] = max(structure["total_cols"], sheet_info["max_col"])
            
            wb.close()
            return structure
            
        except Exception as e:
            logger.error(f"Error analyzing Excel structure for {file_path}: {e}")
            return {}

    def extract_template_config(self, file_path: Path, structure: Dict[str, Any]) -> ExcelTemplateConfig:
        """Extract template configuration from Excel file"""
        try:
            wb = load_workbook(file_path, read_only=True)
            
            # Default config
            config = ExcelTemplateConfig(
                name=file_path.stem,
                description=f"Template migrado desde {file_path.name}",
                category="migrado"
            )
            
            # Try to extract configuration from first sheet
            if structure.get("sheets"):
                first_sheet = structure["sheets"][0]
                ws = wb[first_sheet["name"]]
                
                # Look for configuration in first few rows
                config_found = False
                for row in range(1, min(10, ws.max_row)):
                    for col in range(1, min(10, ws.max_column)):
                        cell = ws.cell(row=row, column=col)
                        if cell.value and isinstance(cell.value, str):
                            value = cell.value.strip().lower()
                            
                            # Extract template name
                            if "nombre" in value or "name" in value:
                                name_cell = ws.cell(row=row, column=col + 1)
                                if name_cell.value:
                                    config.name = str(name_cell.value).strip()
                                    config_found = True
                            
                            # Extract description
                            elif "descripción" in value or "description" in value:
                                desc_cell = ws.cell(row=row, column=col + 1)
                                if desc_cell.value:
                                    config.description = str(desc_cell.value).strip()
                            
                            # Extract category
                            elif "categoría" in value or "category" in value:
                                cat_cell = ws.cell(row=row, column=col + 1)
                                if cat_cell.value:
                                    config.category = str(cat_cell.value).strip()
                            
                            # Extract days per week
                            elif "días" in value or "days" in value:
                                days_cell = ws.cell(row=row, column=col + 1)
                                if days_cell.value and str(days_cell.value).isdigit():
                                    config.days_per_week = int(str(days_cell.value))
                
                # Extract styling information
                config.styling_config = self._extract_styling_config(ws)
                
                # Extract layout information
                config.layout_config = self._extract_layout_config(ws, structure)
                
                # Extract exercise configuration
                config.exercise_config = self._extract_exercise_config(ws, structure)
            
            wb.close()
            return config
            
        except Exception as e:
            logger.error(f"Error extracting template config from {file_path}: {e}")
            return ExcelTemplateConfig(
                name=file_path.stem,
                description=f"Error al migrar: {str(e)}",
                category="error"
            )

    def _extract_styling_config(self, worksheet) -> Dict[str, Any]:
        """Extract styling configuration from worksheet"""
        styling = {
            "fonts": [],
            "colors": [],
            "borders": [],
            "alignments": []
        }
        
        try:
            # Sample cells to extract styling
            sample_cells = []
            for row in range(1, min(20, worksheet.max_row)):
                for col in range(1, min(10, worksheet.max_column)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value:
                        sample_cells.append(cell)
                        if len(sample_cells) >= 10:  # Sample 10 cells
                            break
                if len(sample_cells) >= 10:
                    break
            
            for cell in sample_cells:
                if cell.font:
                    font_info = {
                        "name": cell.font.name,
                        "size": cell.font.size,
                        "bold": cell.font.bold,
                        "italic": cell.font.italic,
                        "color": cell.font.color.rgb if cell.font.color else None
                    }
                    if font_info not in styling["fonts"]:
                        styling["fonts"].append(font_info)
                
                if cell.fill:
                    fill_info = {
                        "type": "solid" if cell.fill.fill_type == "solid" else "pattern",
                        "color": cell.fill.start_color.rgb if cell.fill.start_color else None
                    }
                    if fill_info not in styling["colors"]:
                        styling["colors"].append(fill_info)
        
        except Exception as e:
            logger.warning(f"Error extracting styling config: {e}")
        
        return styling

    def _extract_layout_config(self, worksheet, structure: Dict[str, Any]) -> Dict[str, Any]:
        """Extract layout configuration from worksheet"""
        layout = {
            "orientation": "portrait",
            "page_size": "A4",
            "margins": {
                "top": 20,
                "right": 20,
                "bottom": 20,
                "left": 20
            },
            "column_widths": [],
            "row_heights": [],
            "merged_cells": []
        }
        
        try:
            # Extract column widths
            for col in range(1, min(worksheet.max_column + 1, 27)):  # A-Z
                column_letter = get_column_letter(col)
                width = worksheet.column_dimensions[column_letter].width
                if width:
                    layout["column_widths"].append({
                        "column": column_letter,
                        "width": width
                    })
            
            # Extract row heights
            for row in range(1, min(worksheet.max_row + 1, 101)):  # First 100 rows
                if row in worksheet.row_dimensions:
                    height = worksheet.row_dimensions[row].height
                    if height:
                        layout["row_heights"].append({
                            "row": row,
                            "height": height
                        })
            
            # Extract merged cells
            for merged_range in worksheet.merged_cells.ranges:
                layout["merged_cells"].append(str(merged_range))
        
        except Exception as e:
            logger.warning(f"Error extracting layout config: {e}")
        
        return layout

    def _extract_exercise_config(self, worksheet, structure: Dict[str, Any]) -> Dict[str, Any]:
        """Extract exercise configuration from worksheet"""
        exercise_config = {
            "columns": [],
            "data_start": None,
            "data_end": None,
            "total_exercises": 0
        }
        
        try:
            # Look for exercise table structure
            headers_found = False
            for row in range(1, min(20, worksheet.max_row)):
                for col in range(1, min(10, worksheet.max_column)):
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value and isinstance(cell.value, str):
                        value = cell.value.strip().lower()
                        
                        # Common exercise table headers
                        if any(header in value for header in [
                            "ejercicio", "exercise", "series", "sets", 
                            "repeticiones", "reps", "descanso", "rest"
                        ]):
                            headers_found = True
                            exercise_config["data_start"] = row + 1
                            
                            # Extract column headers
                            header_col = col
                            while header_col <= worksheet.max_column:
                                header_cell = worksheet.cell(row=row, column=header_col)
                                if header_cell.value:
                                    exercise_config["columns"].append({
                                        "column": get_column_letter(header_col),
                                        "header": str(header_cell.value).strip(),
                                        "type": self._detect_column_type(str(header_cell.value).strip())
                                    })
                                header_col += 1
                            
                            break
                
                if headers_found:
                    break
            
            # Find data end
            if exercise_config["data_start"]:
                for row in range(exercise_config["data_start"], worksheet.max_row + 1):
                    has_data = False
                    for col in range(1, len(exercise_config["columns"]) + 1):
                        cell = worksheet.cell(row=row, column=col)
                        if cell.value and str(cell.value).strip():
                            has_data = True
                            break
                    
                    if has_data:
                        exercise_config["data_end"] = row
                    else:
                        break
                
                if exercise_config["data_end"]:
                    exercise_config["total_exercises"] = (
                        exercise_config["data_end"] - exercise_config["data_start"] + 1
                    )
        
        except Exception as e:
            logger.warning(f"Error extracting exercise config: {e}")
        
        return exercise_config

    def _detect_column_type(self, header: str) -> str:
        """Detect column type based on header"""
        header_lower = header.lower()
        
        if any(word in header_lower for word in ["ejercicio", "exercise"]):
            return "exercise_name"
        elif any(word in header_lower for word in ["series", "sets"]):
            return "sets"
        elif any(word in header_lower for word in ["repeticiones", "reps", "rep"]):
            return "reps"
        elif any(word in header_lower for word in ["descanso", "rest"]):
            return "rest"
        elif any(word in header_lower for word in ["peso", "weight", "kg"]):
            return "weight"
        elif any(word in header_lower for word in ["notas", "notes", "comentarios"]):
            return "notes"
        else:
            return "text"

    def extract_exercises(self, file_path: Path, config: ExcelTemplateConfig) -> List[ExerciseMapping]:
        """Extract exercises from Excel template"""
        exercises = []
        
        try:
            wb = load_workbook(file_path, read_only=True)
            
            if config.exercise_config and config.exercise_config.get("data_start"):
                # Use detected exercise configuration
                start_row = config.exercise_config["data_start"]
                end_row = config.exercise_config.get("data_end", start_row)
                columns = config.exercise_config["columns"]
                
                # Find the sheet with exercise data
                for sheet_name in wb.sheetnames:
                    ws = wb[sheet_name]
                    
                    for row in range(start_row, end_row + 1):
                        exercise_data = {}
                        
                        for col_info in columns:
                            col_letter = col_info["column"]
                            col_num = ord(col_letter.upper()) - ord('A') + 1
                            cell = ws.cell(row=row, column=col_num)
                            
                            if cell.value:
                                exercise_data[col_info["type"]] = str(cell.value).strip()
                        
                        # Create exercise mapping if we have essential data
                        if exercise_data.get("exercise_name"):
                            exercise = ExerciseMapping(
                                name=exercise_data["exercise_name"],
                                sets=int(exercise_data.get("sets", 1)),
                                reps=exercise_data.get("reps", "10"),
                                rest=exercise_data.get("rest"),
                                notes=exercise_data.get("notes"),
                                muscle_group=exercise_data.get("muscle_group"),
                                difficulty=exercise_data.get("difficulty")
                            )
                            exercises.append(exercise)
            
            wb.close()
            
        except Exception as e:
            logger.error(f"Error extracting exercises from {file_path}: {e}")
        
        return exercises

    def convert_to_dynamic_template(self, config: ExcelTemplateConfig, exercises: List[ExerciseMapping]) -> Dict[str, Any]:
        """Convert Excel template to dynamic template format"""
        template_config = self.default_config.copy()
        
        # Update metadata
        template_config["metadata"]["name"] = config.name
        template_config["metadata"]["description"] = config.description
        template_config["metadata"]["tags"] = ["migrado", "excel", config.category]
        
        # Create sections based on exercises
        sections = []
        
        # Header section
        sections.append({
            "id": "header",
            "type": "header",
            "content": {
                "title": "{{gym_name}}",
                "subtitle": "Rutina de Entrenamiento",
                "client_name": "{{client_name}}",
                "trainer_name": "{{trainer_name}}",
                "show_logo": True,
                "show_date": True
            }
        })
        
        # Exercise table section
        if exercises:
            # Group exercises by day if possible
            exercise_groups = self._group_exercises_by_day(exercises)
            
            for day_idx, day_exercises in enumerate(exercise_groups):
                sections.append({
                    "id": f"day_{day_idx + 1}",
                    "type": "exercise_table",
                    "content": {
                        "title": f"Día {day_idx + 1}",
                        "exercises": [
                            {
                                "name": ex.name,
                                "sets": ex.sets,
                                "reps": ex.reps,
                                "rest": ex.rest,
                                "notes": ex.notes,
                                "muscle_group": ex.muscle_group,
                                "difficulty": ex.difficulty
                            }
                            for ex in day_exercises
                        ],
                        "show_day_separator": len(exercise_groups) > 1,
                        "columns": ["ejercicio", "series", "repeticiones", "descanso", "notas"]
                    }
                })
        
        # Footer section
        sections.append({
            "id": "footer",
            "type": "footer",
            "content": {
                "show_signature": True,
                "show_date": True,
                "notes": "Consulta con un profesional antes de comenzar cualquier rutina de ejercicios."
            }
        })
        
        template_config["sections"] = sections
        
        # Apply styling from Excel if available
        if config.styling_config:
            template_config["styling"].update(self._convert_excel_styling(config.styling_config))
        
        # Apply layout from Excel if available
        if config.layout_config:
            template_config["layout"].update(self._convert_excel_layout(config.layout_config))
        
        return template_config

    def _group_exercises_by_day(self, exercises: List[ExerciseMapping]) -> List[List[ExerciseMapping]]:
        """Group exercises by day (simplified logic)"""
        if not exercises:
            return []
        
        # For now, group every 6-8 exercises as a day
        # In a real implementation, this would be more sophisticated
        exercises_per_day = 7
        groups = []
        
        for i in range(0, len(exercises), exercises_per_day):
            groups.append(exercises[i:i + exercises_per_day])
        
        return groups

    def _convert_excel_styling(self, excel_styling: Dict[str, Any]) -> Dict[str, Any]:
        """Convert Excel styling to template styling"""
        styling = {}
        
        try:
            # Extract primary font
            if excel_styling.get("fonts"):
                primary_font = excel_styling["fonts"][0]
                if primary_font.get("name"):
                    styling["font_family"] = primary_font["name"]
                if primary_font.get("size"):
                    styling["font_size"] = int(primary_font["size"])
            
            # Extract colors
            if excel_styling.get("colors"):
                for color in excel_styling["colors"]:
                    if color.get("color"):
                        styling["primary_color"] = f"#{color['color']}"
                        break
        
        except Exception as e:
            logger.warning(f"Error converting Excel styling: {e}")
        
        return styling

    def _convert_excel_layout(self, excel_layout: Dict[str, Any]) -> Dict[str, Any]:
        """Convert Excel layout to template layout"""
        layout = {}
        
        try:
            # Convert column widths to template format
            if excel_layout.get("column_widths"):
                layout["column_widths"] = excel_layout["column_widths"]
            
            # Convert row heights to template format
            if excel_layout.get("row_heights"):
                layout["row_heights"] = excel_layout["row_heights"]
        
        except Exception as e:
            logger.warning(f"Error converting Excel layout: {e}")
        
        return layout

    def save_template(self, template_config: Dict[str, Any], config: ExcelTemplateConfig) -> str:
        """Save migrated template to file"""
        try:
            # Create filename
            safe_name = "".join(c for c in config.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
            filename = f"{safe_name}_migrated.json"
            output_path = self.output_dir / filename
            
            # Save template configuration
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(template_config, f, indent=2, ensure_ascii=False)
            
            # Save migration metadata
            metadata = {
                "original_file": config.name,
                "migration_date": datetime.now().isoformat(),
                "template_config": asdict(config),
                "sections_count": len(template_config.get("sections", [])),
                "variables_count": len(template_config.get("variables", {}))
            }
            
            metadata_path = self.output_dir / f"{safe_name}_metadata.json"
            with open(metadata_path, 'w', encoding='utf-8') as f:
                json.dump(metadata, f, indent=2, ensure_ascii=False)
            
            logger.info(f"Template saved: {output_path}")
            return str(output_path)
            
        except Exception as e:
            logger.error(f"Error saving template: {e}")
            return ""

    def migrate_template(self, file_path: Path) -> Tuple[bool, str]:
        """Migrate a single Excel template"""
        try:
            logger.info(f"Migrating template: {file_path.name}")
            
            # Analyze Excel structure
            structure = self.analyze_excel_structure(file_path)
            if not structure:
                return False, "Could not analyze Excel structure"
            
            # Extract configuration
            config = self.extract_template_config(file_path, structure)
            
            # Extract exercises
            exercises = self.extract_exercises(file_path, config)
            
            # Convert to dynamic template
            template_config = self.convert_to_dynamic_template(config, exercises)
            
            # Save template
            output_path = self.save_template(template_config, config)
            
            if output_path:
                # Log migration
                log_entry = {
                    "file": str(file_path),
                    "status": "success",
                    "output": output_path,
                    "exercises_count": len(exercises),
                    "sections_count": len(template_config.get("sections", [])),
                    "timestamp": datetime.now().isoformat()
                }
                self.migration_log.append(log_entry)
                
                return True, output_path
            else:
                return False, "Failed to save template"
                
        except Exception as e:
            logger.error(f"Error migrating template {file_path}: {e}")
            
            log_entry = {
                "file": str(file_path),
                "status": "error",
                "error": str(e),
                "timestamp": datetime.now().isoformat()
            }
            self.migration_log.append(log_entry)
            
            return False, str(e)

    def migrate_all_templates(self) -> Dict[str, Any]:
        """Migrate all Excel templates in directory"""
        results = {
            "total": 0,
            "successful": 0,
            "failed": 0,
            "migrated_templates": [],
            "errors": []
        }
        
        excel_files = self.scan_excel_templates()
        results["total"] = len(excel_files)
        
        for file_path in excel_files:
            success, result = self.migrate_template(file_path)
            
            if success:
                results["successful"] += 1
                results["migrated_templates"].append(result)
            else:
                results["failed"] += 1
                results["errors"].append({
                    "file": str(file_path),
                    "error": result
                })
        
        # Save migration log
        log_path = self.output_dir / "migration_log.json"
        with open(log_path, 'w', encoding='utf-8') as f:
            json.dump({
                "summary": results,
                "detailed_log": self.migration_log
            }, f, indent=2, ensure_ascii=False)
        
        return results

    def generate_migration_report(self) -> str:
        """Generate detailed migration report"""
        if not self.migration_log:
            return "No migrations performed yet."
        
        report = []
        report.append("# Excel Template Migration Report")
        report.append(f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        report.append("")
        
        # Summary
        successful = len([log for log in self.migration_log if log["status"] == "success"])
        failed = len([log for log in self.migration_log if log["status"] == "error"])
        
        report.append("## Summary")
        report.append(f"- Total templates processed: {len(self.migration_log)}")
        report.append(f"- Successful migrations: {successful}")
        report.append(f"- Failed migrations: {failed}")
        report.append(f"- Success rate: {(successful/len(self.migration_log)*100):.1f}%")
        report.append("")
        
        # Successful migrations
        if successful > 0:
            report.append("## Successful Migrations")
            for log in self.migration_log:
                if log["status"] == "success":
                    report.append(f"- ✅ {Path(log['file']).name}")
                    report.append(f"  - Output: {Path(log['output']).name}")
                    report.append(f"  - Exercises: {log['exercises_count']}")
                    report.append(f"  - Sections: {log['sections_count']}")
                    report.append("")
        
        # Failed migrations
        if failed > 0:
            report.append("## Failed Migrations")
            for log in self.migration_log:
                if log["status"] == "error":
                    report.append(f"- ❌ {Path(log['file']).name}")
                    report.append(f"  - Error: {log['error']}")
                    report.append("")
        
        return "\n".join(report)

def main():
    """Main migration function"""
    print("🚀 Starting Excel Template Migration...")
    
    migrator = ExcelTemplateMigrator()
    
    # Migrate all templates
    results = migrator.migrate_all_templates()
    
    # Print results
    print(f"\n📊 Migration Results:")
    print(f"   Total templates: {results['total']}")
    print(f"   Successful: {results['successful']}")
    print(f"   Failed: {results['failed']}")
    print(f"   Success rate: {(results['successful']/results['total']*100):.1f}%")
    
    # Generate report
    report = migrator.generate_migration_report()
    
    # Save report
    report_path = migrator.output_dir / "migration_report.md"
    with open(report_path, 'w', encoding='utf-8') as f:
        f.write(report)
    
    print(f"\n📄 Migration report saved: {report_path}")
    print(f"📁 Migrated templates saved in: {migrator.output_dir}")
    
    return results

if __name__ == "__main__":
    main()
